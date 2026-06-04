import os
import glob
import msoffcrypto
import openpyxl
import io

# Constantes de MAT
SUCURSALES_ADMIN = ['2043', '2856', '2692', '2511', '313']
PASSWORDS = ['2043', 'ADA181016KD6']

def decrypt_excel(file_path):
    # Intentar desencriptar el archivo si está protegido
    try:
        with open(file_path, "rb") as f:
            office_file = msoffcrypto.OfficeFile(f)
            if not office_file.is_encrypted():
                return file_path
                
            print(f"[{os.path.basename(file_path)}] Está protegido con contraseña. Intentando desencriptar...")
            for pwd in PASSWORDS:
                try:
                    office_file.load_key(password=pwd)
                    decrypted_path = file_path.replace(".xlsx", "_decrypted.xlsx").replace(".xlsm", "_decrypted.xlsm")
                    with open(decrypted_path, "wb") as f_dec:
                        office_file.decrypt(f_dec)
                    print(f"[{os.path.basename(file_path)}] Desencriptado correctamente con contraseña: {pwd}")
                    return decrypted_path
                except Exception as e:
                    continue
            print(f"[{os.path.basename(file_path)}] No se pudo desencriptar con ninguna contraseña.")
            return None
    except Exception as e:
        print(f"[{os.path.basename(file_path)}] Error al leer para desencriptar: {e}")
        return None

def clean_excel(file_path, is_comparativo=False):
    print(f"Limpiando {os.path.basename(file_path)}...")
    
    if file_path.lower().endswith('.xls'):
        # Usar pandas para limpiar .xls antiguos y convertirlos a .xlsx
        import pandas as pd
        try:
            df = pd.read_excel(file_path, engine='xlrd')
            # Buscar fila de encabezados
            header_idx = -1
            mat_col = None
            
            # Buscar en las primeras 20 filas
            for i, row in df.head(30).iterrows():
                row_str = row.astype(str).str.lower().str.strip()
                if any(x in ['asesor', 'mat', 'mat / unidad', 'nombre del asesor'] for x in row_str):
                    header_idx = i
                    # Buscar columna MAT
                    for col_name, val in row.items():
                        if str(val).strip().lower() in ['mat', 'mat / unidad', 'suc', 'sucursal', 'matriz']:
                            mat_col = col_name
                            break
                    break
            
            if mat_col is None:
                if "proactivos" in file_path.lower():
                    mat_col = df.columns[1] # col B o C
                elif "asesores sin" in file_path.lower():
                    mat_col = df.columns[2] # col C
                    
            if header_idx != -1 and mat_col is not None:
                # Filtrar: mantener filas hasta el header, y luego las que mat_col esté en SUCURSALES_ADMIN
                header_data = df.iloc[:header_idx+1]
                data_rows = df.iloc[header_idx+1:]
                
                # Filtrar data_rows
                filtered_data = data_rows[data_rows[mat_col].astype(str).str.strip().isin(SUCURSALES_ADMIN) | data_rows[mat_col].astype(str).str.strip().str.lower().str.contains('total', na=False)]
                
                cleaned_df = pd.concat([header_data, filtered_data])
                new_path = file_path + "x" # Convertir a .xlsx
                cleaned_df.to_excel(new_path, index=False, header=False)
                os.remove(file_path)
                print(f"[{os.path.basename(file_path)}] Convertido y limpio (Guardado como xlsx)!")
            else:
                print(f"No se pudo limpiar {file_path} con pandas.")
        except Exception as e:
            print(f"Error limpiando xls con pandas: {e}")
        return

    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    if is_comparativo:
        if 'asesores' in wb.sheetnames:
            ws = wb['asesores']
        else:
            ws = wb.worksheets[1] if len(wb.worksheets) > 1 else wb.worksheets[0]
    else:
        ws = wb.worksheets[0]

    # Identificar fila de encabezados (buscando "Asesor" o "MAT")
    header_row_idx = -1
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
        if row:
            if any(str(cell).strip().lower() in ['asesor', 'mat', 'mat / unidad', 'nombre del asesor'] for cell in row if cell is not None):
                header_row_idx = row_idx
                break

    if header_row_idx == -1:
        print(f"No se encontraron encabezados en {os.path.basename(file_path)}. Omitiendo limpieza.")
        return

    # Buscar índice de columna de SUCURSAL / MAT
    mat_col_idx = -1
    for col_idx, cell in enumerate(ws[header_row_idx], start=1):
        val = str(cell.value).strip().lower() if cell.value else ""
        if val in ['mat', 'mat / unidad', 'suc', 'sucursal', 'matriz']:
            mat_col_idx = col_idx
            break

    if mat_col_idx == -1:
        # En proactivos y asesores sin emision a veces no dice SUCURSAL sino que la sucursal está en la col 2 o 3 (0-indexed)
        # Haremos hardcode temporal si no se encuentra
        if "proactivos" in file_path.lower():
            mat_col_idx = 2 # usualmente col 2
        elif "asesores sin" in file_path.lower():
            mat_col_idx = 2 # usualmente col 2 o 3
        elif "comparativo" in file_path.lower():
            mat_col_idx = 4 # en comparativo de vida (D o E)
        else:
            print(f"No se encontró columna MAT en {os.path.basename(file_path)}. Omitiendo limpieza.")
            return

    rows_to_delete = []
    # Revisar desde la fila siguiente al encabezado hasta el final
    for row_idx in range(ws.max_row, header_row_idx, -1):
        cell_val = ws.cell(row=row_idx, column=mat_col_idx).value
        if cell_val is not None:
            val_str = str(cell_val).strip()
            if 'total' in val_str.lower():
                continue
            if val_str not in SUCURSALES_ADMIN:
                rows_to_delete.append(row_idx)

    # Eliminar filas de abajo hacia arriba para no alterar los índices
    print(f"Se eliminarán {len(rows_to_delete)} filas de otras promotorías...")
    for r in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(r, 1)

    wb.save(file_path)
    print(f"[{os.path.basename(file_path)}] Guardado y limpio!")

def main():
    tmp_dir = os.path.join(os.path.dirname(__file__), '..', 'tmp_admin')
    if not os.path.exists(tmp_dir):
        print("No se encontró la carpeta tmp_admin.")
        return

    excel_files = glob.glob(os.path.join(tmp_dir, '*.xls*'))
    for f in excel_files:
        if "decrypted" in f:
            continue
            
        fname_lower = os.path.basename(f).lower()
        if "proactivo" not in fname_lower and "sin emision" not in fname_lower and "comparativo vida" not in fname_lower:
            print(f"Omitiendo {os.path.basename(f)} porque no es un reporte administrativo objetivo.")
            continue
            
        is_cv = "comparativo" in fname_lower
        
        # Desencriptar si es necesario
        working_file = decrypt_excel(f)
        if not working_file:
            print(f"Omitiendo {f} porque no se pudo desencriptar (o falló lectura).")
            continue
            
        clean_excel(working_file, is_comparativo=is_cv)
        
        # Reemplazar el archivo original con el limpio si se desencriptó
        if working_file != f:
            os.replace(working_file, f)

if __name__ == '__main__':
    main()
