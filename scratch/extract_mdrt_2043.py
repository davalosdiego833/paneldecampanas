import openpyxl

wb = openpyxl.load_workbook('mdrt/MDRT Asesores.xlsm', data_only=True)
sheet = wb['MDRT']

rows_2043 = []
for r in range(17, sheet.max_row + 1):
    oficina = sheet.cell(row=r, column=3).value
    suc = sheet.cell(row=r, column=5).value
    promotor = sheet.cell(row=r, column=6).value
    asesor = sheet.cell(row=r, column=8).value
    total_prima = sheet.cell(row=r, column=23).value
    camino_prima = sheet.cell(row=r, column=38).value
    
    if asesor and total_prima is not None:
        try:
            val = float(total_prima)
            if val > 0 and (str(suc) == '2043' or 'AMBRIZ' in str(promotor).upper() or 'AMBRIZ' in str(oficina).upper()):
                rows_2043.append({
                    'oficina': str(oficina or '').strip(),
                    'promotor': str(promotor or '').strip(),
                    'asesor': str(asesor or '').strip(),
                    'total_prima': val,
                    'camino_prima': str(camino_prima or '').strip()
                })
        except (ValueError, TypeError):
            pass

rows_2043.sort(key=lambda x: x['total_prima'], reverse=True)

print("--- TOP 10 PROMOTORÍA 2043 (AMBRIZ & DÁVALOS) POR CAMINO DE PRIMAS ---")
for idx, item in enumerate(rows_2043[:10], start=1):
    print(f"{idx}. OFICINA: {item['oficina']} | PROMOTOR: {item['promotor']} | ASESOR: {item['asesor']} | TOTAL PRIMA: ${item['total_prima']:,.2f}")
