import openpyxl

wb = openpyxl.load_workbook('mdrt/MDRT Asesores.xlsm', data_only=True)
sheet = wb['MDRT']

rows_data = []
for r in range(17, sheet.max_row + 1):
    oficina = sheet.cell(row=r, column=3).value
    promotor = sheet.cell(row=r, column=6).value
    asesor = sheet.cell(row=r, column=8).value
    total_prima = sheet.cell(row=r, column=23).value
    camino_prima = sheet.cell(row=r, column=38).value
    
    if asesor and total_prima is not None:
        try:
            val = float(total_prima)
            if val > 0:
                rows_data.append({
                    'oficina': str(oficina or '').strip(),
                    'promotor': str(promotor or '').strip(),
                    'asesor': str(asesor or '').strip(),
                    'total_prima': val,
                    'camino_prima': str(camino_prima or '').strip()
                })
        except (ValueError, TypeError):
            pass

rows_data.sort(key=lambda x: x['total_prima'], reverse=True)

print("--- TOP 15 GLOBAL POR CAMINO DE PRIMAS ---")
for idx, item in enumerate(rows_data[:15], start=1):
    print(f"{idx}. OFICINA: {item['oficina']} | PROMOTOR: {item['promotor']} | ASESOR: {item['asesor']} | TOTAL PRIMA: ${item['total_prima']:,.2f} | CAMINO PRIMA: {item['camino_prima']}")
